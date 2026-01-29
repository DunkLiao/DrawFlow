"""
Excel 生成模塊：將圖表保存到 Excel 檔案
"""

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
import logging
import tempfile
import os
from PIL import Image as PILImage

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Excel 檔案生成器"""

    def __init__(self, dpi=300):
        """
        初始化 Excel 生成器

        Args:
            dpi (int): 圖表 DPI 解析度
        """
        self.dpi = dpi
        self.temp_images = []  # 暫存圖片路徑列表

    def save_excel(self, figures, output_path, stats, title="里程碑時間線"):
        """
        將圖表和統計信息保存為 Excel 檔案

        Args:
            figures (list): matplotlib Figure 物件列表
            output_path (str): 輸出檔案路徑
            stats (dict): 統計信息字典
            title (str): 報告標題

        Returns:
            Path: 輸出檔案路徑
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            # 建立工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "時間線"

            # 設定欄寬和行高
            ws.column_dimensions['A'].width = 40
            ws.row_dimensions[1].height = 30

            # 標題
            ws['A1'] = title
            ws['A1'].font = Font(name='SimHei', size=20,
                                 bold=True, color='FFFFFF')
            ws['A1'].fill = PatternFill(
                start_color='4472C4', end_color='4472C4', fill_type='solid')
            ws['A1'].alignment = Alignment(
                horizontal='center', vertical='center')
            ws.merge_cells('A1:D1')

            # 統計信息
            current_row = 3
            stats_style = Font(name='SimHei', size=12)

            ws[f'A{current_row}'] = "統計信息"
            ws[f'A{current_row}'].font = Font(
                name='SimHei', size=14, bold=True)
            current_row += 1

            # 統計數據表格
            stat_items = [
                ('里程碑總數', f"{stats['total_milestones']}"),
                ('開始日期', f"{stats['start_date'].strftime('%Y-%m-%d')}"),
                ('結束日期', f"{stats['end_date'].strftime('%Y-%m-%d')}"),
                ('時間跨度', f"{stats['total_days']} 天"),
                ('里程碑密度', f"{stats['milestone_density']:.2f} 個/月"),
            ]

            for label, value in stat_items:
                ws[f'A{current_row}'] = label
                ws[f'B{current_row}'] = value
                ws[f'A{current_row}'].font = stats_style
                ws[f'B{current_row}'].font = stats_style
                ws[f'A{current_row}'].fill = PatternFill(
                    start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                ws[f'B{current_row}'].fill = PatternFill(
                    start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                current_row += 1

            # 圖表
            current_row += 2
            ws[f'A{current_row}'] = "時間線圖表"
            ws[f'A{current_row}'].font = Font(
                name='SimHei', size=14, bold=True)
            current_row += 1

            # 將每個 Figure 保存為圖片並插入到 Excel
            for idx, fig in enumerate(figures, 1):
                # 暫存圖片（DPI=150 + 增加 Figure 尺寸）
                temp_image_path = self._save_figure_as_image(fig, idx, dpi=150)

                # 設定行高以適應圖片
                ws.row_dimensions[current_row].height = 380  # 約 5 英寸高度

                # 插入圖片
                img = XLImage(temp_image_path)
                img.width = 700  # 像素寬度，約 9.3 英寸
                img.height = 380  # 像素高度
                ws.add_image(img, f'A{current_row}')

                current_row += 21  # 每張圖片佔用約 21 行
                if idx < len(figures):
                    current_row += 2

            # 保存工作簿
            wb.save(str(output_path))

            # 清理暫存圖片
            self._cleanup_temp_images()

            logger.info(f"Excel 生成成功: {output_path}")
            logger.info(f"總共 {len(figures)} 頁")

            return output_path

        except Exception as e:
            self._cleanup_temp_images()
            logger.error(f"Excel 生成失敗: {str(e)}", exc_info=True)
            raise

    def _save_figure_as_image(self, fig, page_num, dpi=150):
        """
        將 matplotlib Figure 保存為臨時圖片

        Args:
            fig (matplotlib.figure.Figure): 圖表物件
            page_num (int): 頁碼
            dpi (int): 圖片解析度

        Returns:
            str: 圖片檔案路徑
        """
        # 建立暫存目錄
        temp_dir = tempfile.gettempdir()
        temp_image_path = os.path.join(
            temp_dir,
            f'timeline_page_{page_num}_{os.getpid()}.png'
        )

        # 先以 150 DPI 保存
        fig.savefig(
            temp_image_path,
            dpi=150,
            bbox_inches='tight',
            facecolor='white',
            edgecolor='none',
            pad_inches=0.1
        )

        # 調整圖片大小以確保符合 PIL 限制
        try:
            img = PILImage.open(temp_image_path)
            # 限制圖片最大尺寸，防止超過 PIL 限制
            max_size = (1400, 900)
            img.thumbnail(max_size, PILImage.Resampling.LANCZOS)
            img.save(temp_image_path, 'PNG', optimize=True)
            img.close()
        except Exception as e:
            logger.warning(f"調整圖片大小失敗: {str(e)}")

        self.temp_images.append(temp_image_path)
        logger.debug(f"暫存圖片: {temp_image_path}")

        return temp_image_path

    def _cleanup_temp_images(self):
        """清理暫存圖片"""
        for temp_image_path in self.temp_images:
            try:
                if os.path.exists(temp_image_path):
                    os.remove(temp_image_path)
                    logger.debug(f"刪除暫存圖片: {temp_image_path}")
            except Exception as e:
                logger.warning(f"無法刪除暫存圖片 {temp_image_path}: {str(e)}")

        self.temp_images.clear()
